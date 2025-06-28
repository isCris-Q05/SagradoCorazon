from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import Usuario, Paciente, Medico, Alergia, Especialidad, MedicoEspecialidad, Enfermedad, Tratamiento,TratamientoEnfermedad, Producto, Cita, Registro, RegistroTratamiento, RegistroProducto, PacienteEnfermedad, PacienteAlergia
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.utils.timezone import now
from citas.utils import admin_medico_required
import pywhatkit as kit
from django.views.decorators.csrf import csrf_exempt
import random
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models.functions import ExtractMonth, ExtractYear
from django.db.models.functions import TruncMonth
# importando date
from datetime import date, datetime



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import pandas as pd
from .forms import UploadExcelForm

from datetime import datetime, timedelta


from django.conf import settings
from twilio.rest import Client


# Create your views here.
def index(request):
    return render(request, "citas/index.html")


def paciente_inicio(request):
    return render(request,'citas2/vista_pacientes/inicio.html')

def paciente_citas(request):
    paciente_id = 5

    proximas_citas = Cita.objects.filter(fecha__gte=now(), id_paciente=paciente_id).order_by('fecha')
    citas_pasadas = Cita.objects.filter(fecha__lt=now(), id_paciente=paciente_id).order_by('-fecha')

    return render(request, 'citas2/vista_pacientes/MisCitas.html', {
        'proximas_citas': proximas_citas,
        'citas_pasadas': citas_pasadas
    })

def paciente_historial(request):
    return render(request,'citas2/vista_pacientes/HistorialMedico.html')


# registrar pacientes
def registrar_paciente(request):
    pacientes = Paciente.objects.all()

    if request.method == "POST":
        username = request.POST.get('username')
        nombre = request.POST.get('nombrePaciente')
        apellido = request.POST.get('apellidoPaciente')
        fecha_nacimiento = request.POST.get('fechaNacimientoPaciente')
        cedula = request.POST.get('cedulaPaciente')
        genero = request.POST.get('generoPaciente')
        telefono = request.POST.get('telefonoPaciente')
        email = request.POST.get('emailPaciente')
        direccion = request.POST.get('direccionPaciente')
        motivo = request.POST.get('motivoConsulta')
        enfermedades = request.POST.getlist('enfermedades')  # Recoger enfermedades como lista
        alergias = request.POST.getlist('alergias')  # Recoger alergias como lista
        contacto_emergencia = request.POST.get('contactoEmergencia')
        telefono_emergencia = request.POST.get('telefonoEmergencia')

        print(f"user: {username}")
        #return HttpResponse("pacientte creado")
        
         # Crear un usuario si es necesario (ejemplo: se puede tener un usuario para el paciente)
        usuario = Usuario.objects.create(
            username = username,
            first_name=nombre,
            last_name=apellido,
            email=email
        )
        
        # Crear el paciente
        paciente = Paciente.objects.create(
            user=usuario,
            genero=genero,
            direccion=direccion,
            motivo=motivo,
            cedula=cedula,
            telefono=telefono,
            fecha_nacimiento=fecha_nacimiento,
            enfermedadess=enfermedades,
            alergiass=alergias,
            contacto_emergencias=contacto_emergencia,
            telefono_emergencia=telefono_emergencia,
            apellido=apellido
        )

        request.session['paciente_nombre'] = f"{usuario.first_name} {usuario.last_name}"
        messages.success(request, 'El paciente se ha registrado correctamente.')
        return redirect('inicio')
    return render(request,"citas2/pacientes.html",{'pacientes':pacientes})

#@admin_medico_required    
def registrar_medico(request):
    if request.method == "POST":
        # Datos del usuario
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password', '')  # Solo una contraseña
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        role = 'Medico'
        genero = request.POST.get('genero', '')

        # Datos del médico
        telefono = request.POST.get('telefono', '')
        is_admin = request.POST.get('is_admin', 'False') == 'True'

        print(f"is_admin: {is_admin}")

        print(f"username: {username}")
        print(f"email: {email}")
        print(f"password: {password}")
        print(f"first_name: {first_name}")
        print(f"last_name: {last_name}")
        print(f"role: {role}")
        print(f"genero: {genero}")
        print(f"telefono: {telefono}")

        #return HttpResponse("adadhadhad creado")

        # Validaciones
        if Usuario.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya está registrado.')
            return redirect('doctores')

        if Usuario.objects.filter(email=email).exists():
            messages.error(request, 'El correo electrónico ya está registrado.')
            return redirect('doctores')

        # Crear el usuario
        user = Usuario.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role=role,
            genero=genero
        )

        # Crear el médico
        medico = Medico.objects.create(
            user=user,
            telefono=telefono,
            es_admin=is_admin  # Asignar el valor de is_admin
        )

        # Guardar especialidades
        especialidades_ids = request.POST.getlist('especialidades')
        for especialidad_id in especialidades_ids:
            especialidad = Especialidad.objects.get(id_especialidad=especialidad_id)
            MedicoEspecialidad.objects.create(
                id_medico=medico,
                id_especialidad=especialidad
            )

        messages.success(request, 'El médico se ha registrado correctamente.')
        return redirect('doctores')

    #return redirect('register')


# login para medicos
def login_medico(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if hasattr(user, 'role') and user.role == 'Medico':
                login(request, user)
                return redirect('inicio')
            else:
                messages.error(request, 'El usuario no es un medico.')
                return redirect('login')
        else:
            messages.error(request, 'El usuario o la contraseña son incorrectos.')
            return redirect('login')

import json
from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.contrib import messages
from .models import Usuario  # Asegúrate de importar tu modelo Usuario

def forgot_password_medico(request):
    if request.method == "POST":
        email = request.POST['email_recovery']
        print(f"email: {email}")

        # Validamos que el email exista
        if Usuario.objects.filter(email=email).exists():
            user = Usuario.objects.get(email=email)
            if user.role == 'Medico':
                # Generar OTP
                otp = random.randint(100000, 999999)
                request.session['otp'] = otp
                request.session['email'] = email

                # Crear el mensaje del correo
                subject = 'Recuperación de contraseña'
                message = f'Su código de recuperación es: {otp}'
                from_email = 'cristopherquintana2725@gmail.com'  # Remitente (tu dirección de Gmail)
                recipient_list = [email]  # Destinatario

                # Enviar el correo
                try:
                    email_message = EmailMessage(subject, message, from_email, recipient_list)
                    email_message.send(fail_silently=False)
                    return JsonResponse({'status': 'success', 'message': 'Se ha enviado un correo para restablecer la contraseña.'})
                except Exception as e:
                    print(f"Error al enviar el correo: {e}")
                    return JsonResponse({'status': 'error', 'message': 'Error al enviar el correo. Inténtelo de nuevo más tarde.'}, status=500)
            else:
                return JsonResponse({'status': 'error', 'message': 'El usuario no es un médico.'}, status=400)
        else:
            return JsonResponse({'status': 'error', 'message': 'El correo no existe.'}, status=400)


from django.http import JsonResponse
from django.contrib.auth.hashers import make_password

def validate_otp(request):
    if request.method == "POST":
        otp_entered = request.POST.get('otp')
        otp_session = request.session.get('otp')

        if otp_entered and otp_session and int(otp_entered) == otp_session:
            # OTP correcto, permitir cambiar la contraseña
            return JsonResponse({'status': 'success', 'message': 'OTP correcto.', 'allow_password_change': True})
        else:
            return JsonResponse({'status': 'error', 'message': 'OTP incorrecto.'}, status=400)

def change_password(request):
    if request.method == "POST":
        new_password = request.POST.get('new_password')
        email = request.session.get('email')

        if email and new_password:
            try:
                user = Usuario.objects.get(email=email)
                user.password = make_password(new_password)  # Hashear la nueva contraseña
                user.save()
                return JsonResponse({'status': 'success', 'message': 'Contraseña cambiada correctamente.'})
            except Usuario.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Usuario no encontrado.'}, status=400)
        else:
            return JsonResponse({'status': 'error', 'message': 'Datos incompletos.'}, status=400)

@csrf_exempt
def send_reminder(request):
    if request.method == "POST":
        phone_number = request.POST.get('phone_number')
        message = request.POST.get('message')

        print(f"phone_number: {phone_number}")

        try:
            # Enviar mensaje de recordatorio
            kit.sendwhatmsg_instantly(phone_number, message)
            return JsonResponse({'status': 'success', 'message': 'Recordatorio enviado correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error al enviar el recordatorio: {str(e)}'}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

# login para pacientes
def login_paciente(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']

        print(f"username: {username}, password: {password}")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if  hasattr(user, 'role') and user.role == 'Paciente':
                login(request, user)
                return redirect('dashboard_paciente')
            else:
                messages.error(request, 'El usuario no es un paciente.')
                return redirect('login')
        else:
            messages.error(request, 'El usuario o la contraseña son incorrectos.')
            return redirect('login')
    
def login_usuario(request):
    return render(request, 'citas2/login.html')

def register(request):
    return render(request, 'citas2/register.html')

from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Q

def actualizar_estado_citas():
    """
    Actualiza citas pendientes a "No asistió" cuando han pasado 15 minutos 
    de la hora programada sin ser marcadas como "Finalizada"
    """
    # Obtener la hora actual en la zona horaria de Managua
    ahora_managua = timezone.localtime(timezone.now())
    umbral_tiempo = ahora_managua - timedelta(minutes=15)
    
    # Debug información
    print(f"[DEBUG] Hora actual en Managua: {ahora_managua}")
    print(f"[DEBUG] Umbral de tiempo (hora actual -15min): {umbral_tiempo}")
    
    # Obtener citas pendientes que deberían marcarse como no asistidas
    citas_no_asistidas = Cita.objects.filter(
        estado=Cita.PENDIENTE  # Solo citas pendientes
    ).filter(
        Q(fecha__lt=ahora_managua.date()) |  # Citas de días anteriores
        Q(
            fecha=ahora_managua.date(),
            hora__lte=umbral_tiempo.time()  # Citas de hoy con hora pasada +15min
        )
    )
    
    # Debug detallado
    print(f"[DEBUG] Citas pendientes encontradas para marcar como no asistidas: {citas_no_asistidas.count()}")
    
    for cita in citas_no_asistidas[:5]:  # Mostrar solo las primeras 5 para debug
        # Convertir a datetime con zona horaria para la comparación
        cita_datetime = timezone.make_aware(
            datetime.combine(cita.fecha, cita.hora),
            timezone.get_current_timezone()
        )
        print(f"[DEBUG] Cita ID {cita.id_cita}: Programada {cita.fecha} {cita.hora}")
        print(f"[DEBUG] Tiempo transcurrido: {ahora_managua - cita_datetime}")
    
    # Actualizar estado a "No asistió"
    actualizadas = citas_no_asistidas.update(estado=Cita.NO_ASISTIO)
    print(f"[DEBUG] Total citas actualizadas a 'No asistió': {actualizadas}")
    
    return actualizadas

@login_required
def inicio(request):
    # Procesamiento inicial para la página (sin cambios)
    citas_actualizadas = actualizar_estado_citas()
    print(f"Citas actualizadas: {citas_actualizadas}")

    citas_list = Cita.objects.all().order_by('-fecha', '-hora')
    registros_list = Registro.objects.select_related('id_enfermedad', 'id_cita').prefetch_related('tratamientos', 'registroproducto_set').order_by('-id_cita')
    
    citas_paginator = Paginator(citas_list, 10)
    registros_paginator = Paginator(registros_list, 10)

    enfermedades = Enfermedad.objects.all()
    productos = Producto.objects.all()

    page_number = request.GET.get('page', 1)

    try:
        citas = citas_paginator.get_page(page_number)
        registros = registros_paginator.get_page(page_number)
    except PageNotAnInteger:
        citas = citas_paginator.page(1)
        registros = registros_paginator.page(1)
    except EmptyPage:
        citas = citas_paginator.page(citas_paginator.num_pages)
        registros = registros_paginator.page(registros_paginator.num_pages)

    # Manejo del formulario de Excel (AJAX)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        response_data = {'status': 'success', 'message': '', 'created': 0, 'skipped': 0, 'errors': []}
        print(f"excel_file: {excel_file}")
        try:
            xls = pd.ExcelFile(excel_file)

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
                # Convertir nombres de columnas a minúsculas y sin espacios
                df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]
            
                # Columnas requeridas con diferentes posibles nombres
                column_mapping = {
                    'cedula': ['cedula', 'cédula', 'cédula_paciente'],
                    'id_cita': ['id_cita', 'idcita', 'cita', 'n°_cita'],
                    'medico_user': ['medico_user', 'medicouser', 'username_medico', 'usuario_medico'],
                    'fecha': ['fecha', 'fecha_cita'],
                    'hora': ['hora', 'hora_cita', 'hora_cita'],
                    'id_especialidad': ['id_especialidad', 'especialidad'],
                    'estado': ['estado', 'estado_cita']
                }
            
                # Verificar columnas
                missing_cols = []
                for key, aliases in column_mapping.items():
                    if not any(alias in df.columns for alias in aliases):
                        missing_cols.append(key)
            
                if missing_cols:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'Columnas requeridas faltantes: {", ".join(missing_cols)}'
                    }, status=400)
            
                # Normalizar nombres de columnas
                df = df.rename(columns={
                    next(alias for alias in column_mapping['cedula'] if alias in df.columns): 'cedula',
                    next(alias for alias in column_mapping['id_cita'] if alias in df.columns): 'id_cita',
                    next(alias for alias in column_mapping['medico_user'] if alias in df.columns): 'medico_user',
                    next(alias for alias in column_mapping['fecha'] if alias in df.columns): 'fecha',
                    next(alias for alias in column_mapping['hora'] if alias in df.columns): 'hora',
                    next(alias for alias in column_mapping['estado'] if alias in df.columns): 'estado',
                    next(alias for alias in column_mapping['id_especialidad'] if alias in df.columns): 'especialidad',  # Cambiado a 'especialidad'
                })
            
                for index, row in df.iterrows():
                    try:
                        # Saltar filas vacías
                        if pd.isna(row['id_cita']):
                            continue
                        
                        # Verificar si la cita ya existe
                        if Cita.objects.filter(id_cita=row['id_cita']).exists():
                            response_data['skipped'] += 1
                            continue
                        
                        # Buscar paciente por cédula
                        try:
                            paciente = Paciente.objects.get(cedula=row['cedula'])
                        except Paciente.DoesNotExist:
                            print(f"Error 1")
                            response_data['errors'].append(f"Fila {index+1}: Paciente con cédula {row['cedula']} no encontrado")
                            continue
                        
                        # Buscar médico por username
                        try:
                            medico_user = Usuario.objects.get(username=row['medico_user'], role='Medico')
                            medico = Medico.objects.get(user=medico_user)
                        except Usuario.DoesNotExist:
                            print(f"Error 2")
                            response_data['errors'].append(f"Hoja '{sheet_name}', Fila {index+2}: Médico con username {row['medico_user']} no encontrado")                            
                            continue
                        except Medico.DoesNotExist:
                            print(f"Error 3")
                            response_data['errors'].append(f"Hoja '{sheet_name}', Fila {index+2}: Perfil médico no existe para {row['medico_user']}")                            
                            continue
                        
                        # Convertir fecha y hora
                        try:
                            fecha = pd.to_datetime(row['fecha']).date()
                            hora = pd.to_datetime(row['hora']).time()
                        except Exception as e:
                            print(f"Error 4")
                            response_data['errors'].append(f"Hoja '{sheet_name}', Fila {index+2}: Error en fecha/hora - {str(e)}")
                            continue

                        # antes de crear la cita, tomamos la especialidad
                        try:
                            if pd.isna(row['especialidad']):
                                raise ValueError("Especialidad no proporcionada")
                            
                            especialidad_nombre = str(row['especialidad']).strip()

                            # Buscar especialidad
                            especialidad = Especialidad.objects.filter(
                                            nombre__iexact=especialidad_nombre
                                        ).first()

                            if not especialidad:
                                print(f"Error 6: Especialidad '{row['especialidad']}' no encontrada")
                                response_data['errors'].append(
                                    f"Hoja '{sheet_name}', Fila {index+2}: Especialidad '{row['especialidad']}' no encontrada. "
                                    f"Especialidades disponibles: {list(Especialidad.objects.values_list('nombre', flat=True))}"
                                )
                                continue

                        except Exception as e:
                            print(f"Error al procesar especialidad: {str(e)}")
                            response_data['errors'].append(
                                f"Hoja '{sheet_name}', Fila {index+2}: Error en especialidad - {str(e)}"
                            )
                            continue  

                        # Crear la cita
                        Cita.objects.create(
                            id_cita=row['id_cita'],
                            id_paciente=paciente,
                            id_medico=medico,
                            fecha=fecha,
                            hora=hora,
                            id_especialidad=especialidad,
                            estado=row['estado']
                        )
                        response_data['created'] += 1
                        
                    except Exception as e:
                        print(f"Error 5")
                        response_data['errors'].append(f"Hoja '{sheet_name}', Fila {index+2}: Error - {str(e)}")
                        print(f"Error: {str(e)}")
                        continue
            
            # Construir mensaje final
            if response_data['errors']:
                response_data['status'] = 'partial'
            
            response_data['message'] = (
                f"Procesado: {response_data['created']} nuevas citas, "
                f"{response_data['skipped']} existentes, "
                f"{len(response_data['errors'])} errores"
            )
            
            return JsonResponse(response_data)
        
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al procesar el archivo: {str(e)}'
            }, status=400)

    # GET request normal
    paciente_nombre = request.session.pop('paciente_nombre', None)
    print("Cantidad de citas:", citas_list.count())
    cantidad_citas = citas_list.count()
    print(paciente_nombre)
    form = UploadExcelForm()
    return render(request, "citas2/index.html", {
        "paciente_nombre": paciente_nombre,
        "citas": citas,
        "registros": registros,
        "enfermedades": enfermedades,
        "productos": productos,
        "form": form,
        "cantidad_citas": cantidad_citas
    })

def cantidad_total_citas(request):
    citas = Cita.objects.all()
    return JsonResponse({
        "success": True,
        "count": citas.count()
    })

def visualizacion(request):
    return render(request, "citas2/visualizacion.html")

# endpoint o vista que devuelve todas las citas en las que asistieron y cantidad total
# se le pasara el campo fecha, tiene que ser YYYY-MM por motivos estadisticos
# ahora lo haremos con un rango de fecha inicial y final, pero si no se pasa nada, se mostraran todas las citas finalizadas
# formato de fecha: YYYY-MM
def citas_asistio(request, fecha_inicio=None, fecha_fin=None):
    try:
        # Base query
        citas = Cita.objects.filter(estado='finalizada').select_related(
            'id_paciente', 'id_medico', 'id_especialidad'
        )
        
        # Filtro por médico (si se proporciona)
        medico_id = request.GET.get('medico_id')
        medico_username = request.GET.get('medico_username')
        
        if medico_id or medico_username:
            try:
                if medico_id:
                    medico = Medico.objects.get(id=medico_id)
                else:
                    medico = Medico.objects.get(user__username=medico_username)
                
                citas = citas.filter(id_medico=medico)
            except Medico.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "El médico no existe"
                }, status=404)
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "ID de médico inválido"
                }, status=400)
        
        # Validación de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            try:
                # Convertir fechas a objetos datetime
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m')
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m')
                
                # Validar que fecha_inicio <= fecha_fin
                if fecha_inicio_obj > fecha_fin_obj:
                    return JsonResponse({
                        "success": False,
                        "message": "La fecha de inicio no puede ser mayor a la fecha final"
                    }, status=400)
                
                print(f"Filtrando citas entre {fecha_inicio} y {fecha_fin}")
                
                # Aplicar filtro de rango de fechas
                citas = citas.filter(
                    Q(fecha__year__gte=fecha_inicio_obj.year, 
                      fecha__month__gte=fecha_inicio_obj.month) &
                    Q(fecha__year__lte=fecha_fin_obj.year,
                      fecha__month__lte=fecha_fin_obj.month)
                )
                
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "Formato de fecha inválido. Use YYYY-MM para ambas fechas"
                }, status=400)

        # Serialización de datos
        citas_data = []
        for cita in citas:
            citas_data.append({
                'id': cita.id_cita,
                'paciente': {
                    'id': cita.id_paciente.id,
                    'nombre': f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
                    'cedula': cita.id_paciente.cedula,
                    'telefono': cita.id_paciente.telefono
                },
                'medico': {
                    'id': cita.id_medico.id,
                    'nombre': f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
                    'especialidad': cita.id_especialidad.nombre,
                    'username': cita.id_medico.user.username
                },
                'fecha': cita.fecha.strftime('%Y-%m-%d'),
                'hora': cita.hora.strftime('%H:%M'),
                'estado': cita.get_estado_display()
            })

        return JsonResponse({
            "success": True,
            "citas": citas_data,
            "count": len(citas_data),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "medico_filtrado": {
                'id': medico.id if medico_id or medico_username else None,
                'nombre': f"{medico.user.first_name} {medico.user.last_name}" if medico_id or medico_username else None
            }
        })

    except Exception as e:
        print(f"Error en citas_asistio: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

# endpoint o vista que devuelve todas las citas en las que no asistieron y cantidad total
def citas_no_asistio(request, fecha_inicio=None, fecha_fin=None):
    try:
        # Base query
        citas = Cita.objects.filter(estado='no_asistio').select_related(
            'id_paciente', 'id_medico', 'id_especialidad'
        ).order_by('-fecha', '-hora')  # Ordenar por fecha y hora descendente
        
        # Filtro por médico (si se proporciona)
        medico_id = request.GET.get('medico_id')
        medico_username = request.GET.get('medico_username')
        
        if medico_id or medico_username:
            try:
                if medico_id:
                    medico = Medico.objects.get(id=medico_id)
                else:
                    medico = Medico.objects.get(user__username=medico_username)
                
                citas = citas.filter(id_medico=medico)
            except Medico.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "El médico no existe"
                }, status=404)
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "ID de médico inválido"
                }, status=400)
        
        # Validación de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            try:
                # Convertir fechas a objetos datetime
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m')
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m')
                
                # Validar que fecha_inicio <= fecha_fin
                if fecha_inicio_obj > fecha_fin_obj:
                    return JsonResponse({
                        "success": False,
                        "message": "La fecha de inicio no puede ser mayor a la fecha final"
                    }, status=400)
                
                print(f"Filtrando citas no asistidas entre {fecha_inicio} y {fecha_fin}")
                
                # Aplicar filtro de rango de fechas
                citas = citas.filter(
                    Q(fecha__year__gte=fecha_inicio_obj.year, 
                      fecha__month__gte=fecha_inicio_obj.month) &
                    Q(fecha__year__lte=fecha_fin_obj.year,
                      fecha__month__lte=fecha_fin_obj.month)
                )
                
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "Formato de fecha inválido. Use YYYY-MM para ambas fechas"
                }, status=400)

        # Serialización de datos
        citas_data = []
        for cita in citas:
            citas_data.append({
                'id': cita.id_cita,
                'paciente': {
                    'id': cita.id_paciente.id,
                    'nombre': f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
                    'cedula': cita.id_paciente.cedula,
                    'telefono': cita.id_paciente.telefono,
                    'email': cita.id_paciente.user.email
                },
                'medico': {
                    'id': cita.id_medico.id,
                    'nombre': f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
                    'especialidad': cita.id_especialidad.nombre,
                    'username': cita.id_medico.user.username,
                    'telefono': cita.id_medico.telefono
                },
                'fecha': cita.fecha.strftime('%Y-%m-%d'),
                'hora': cita.hora.strftime('%H:%M'),
                'estado': cita.get_estado_display(),
                'motivo_no_asistio': "Por registrar",  # Campo para registrar el motivo
                'reagendada': False  # Campo para tracking de reagendación
            })

        return JsonResponse({
            "success": True,
            "citas": citas_data,
            "count": len(citas_data),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "medico_filtrado": {
                'id': medico.id if medico_id or medico_username else None,
                'nombre': f"{medico.user.first_name} {medico.user.last_name}" if medico_id or medico_username else None
            }
        })

    except Exception as e:
        print(f"Error en citas_no_asistio: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

# endpoint o vista que devuelve todas las citas pendientes y cantidad total
def citas_pendientes(request, fecha_inicio=None, fecha_fin=None):
    try:
        # Base query
        citas = Cita.objects.filter(estado='pendiente').select_related(
            'id_paciente', 'id_medico', 'id_especialidad'
        ).order_by('-fecha', '-hora')  # Ordenar por fecha y hora ascendente
        
        # Filtro por médico (si se proporciona)
        medico_id = request.GET.get('medico_id')
        medico_username = request.GET.get('medico_username')
        medico = None
        
        if medico_id or medico_username:
            try:
                if medico_id:
                    medico = Medico.objects.get(id=medico_id)
                else:
                    medico = Medico.objects.get(user__username=medico_username)
                
                citas = citas.filter(id_medico=medico)
            except Medico.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "El médico no existe"
                }, status=404)
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "ID de médico inválido"
                }, status=400)
        
        # Validación de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            try:
                # Convertir fechas a objetos datetime
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m')
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m')
                
                # Validar que fecha_inicio <= fecha_fin
                if fecha_inicio_obj > fecha_fin_obj:
                    return JsonResponse({
                        "success": False,
                        "message": "La fecha de inicio no puede ser mayor a la fecha final"
                    }, status=400)
                
                print(f"Filtrando citas pendientes entre {fecha_inicio} y {fecha_fin}")
                
                # Aplicar filtro de rango de fechas
                citas = citas.filter(
                    Q(fecha__year__gte=fecha_inicio_obj.year, 
                      fecha__month__gte=fecha_inicio_obj.month) &
                    Q(fecha__year__lte=fecha_fin_obj.year,
                      fecha__month__lte=fecha_fin_obj.month)
                )
                
            except ValueError:
                return JsonResponse({
                    "success": False,
                    "message": "Formato de fecha inválido. Use YYYY-MM para ambas fechas"
                }, status=400)
        else:
            print("Mostrando todas las citas pendientes (sin filtro de fecha)")

        # Serialización de datos
        citas_data = []
        for cita in citas:
            citas_data.append({
                'id': cita.id_cita,
                'paciente': {
                    'id': cita.id_paciente.id,
                    'nombre': f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
                    'cedula': cita.id_paciente.cedula,
                    'telefono': cita.id_paciente.telefono,
                    'email': cita.id_paciente.user.email
                },
                'medico': {
                    'id': cita.id_medico.id,
                    'nombre': f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
                    'especialidad': cita.id_especialidad.nombre,
                    'username': cita.id_medico.user.username,
                    'telefono': cita.id_medico.telefono
                },
                'fecha': cita.fecha.strftime('%Y-%m-%d'),
                'hora': cita.hora.strftime('%H:%M'),
                'estado': cita.get_estado_display(),
                'notificado': False  # Campo para tracking de notificaciones
            })

        return JsonResponse({
            "success": True,
            "citas": citas_data,
            "count": len(citas_data),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "medico_filtrado": {
                'id': medico.id if medico_id or medico_username else None,
                'nombre': f"{medico.user.first_name} {medico.user.last_name}" if medico_id or medico_username else None,
                'username': medico.user.username if medico_id or medico_username else None
            },
            "proximas_citas": len([c for c in citas_data if c['fecha'] >= datetime.now().strftime('%Y-%m-%d')])
        })

    except Exception as e:
        print(f"Error en citas_pendientes: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

# Todas las enfermedades, y la cantidad de pacientes que afectan
def enfermedades_cantidad_pacientes(request):
    try:
        # Obtener todas las enfermedades con conteo de pacientes
        enfermedades = Enfermedad.objects.annotate(
            # Pacientes con la enfermedad registrada directamente
            pacientes_directos=Count('pacienteenfermedad__id_paciente', distinct=True),
            # Pacientes con la enfermedad mencionada en registros médicos
            pacientes_en_registros=Count('registro__id_cita__id_paciente', distinct=True)
        ).annotate(
            # Total de pacientes únicos (suma de ambos tipos)
            total_pacientes=F('pacientes_directos') + F('pacientes_en_registros')
        )
        
        # Preparar la respuesta
        response_data = {
            "success": True,
            "enfermedades": [
                {
                    "id": enfermedad.id_enfermedad,
                    "nombre": enfermedad.nombre,
                    "pacientes_directos": enfermedad.pacientes_directos,
                    "pacientes_en_registros": enfermedad.pacientes_en_registros,
                    "total_pacientes": enfermedad.total_pacientes
                }
                for enfermedad in enfermedades
            ]
        }
        
        return JsonResponse(response_data)
    
    except Exception as e:
        print(f"Error en enfermedades_cantidad_pacientes: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)
    
def buscar_enfermedades(request):
    term = request.GET.get('term', '').lower()
    enfermedades = Enfermedad.objects.filter(
        Q(nombre__icontains=term) | Q(descripcion__icontains=term)
    )[:10]
    
    results = []
    for enfermedad in enfermedades:
        # Resaltar coincidencias en el nombre
        nombre = enfermedad.nombre
        idx = nombre.lower().find(term)
        if idx >= 0:
            highlighted = (
                nombre[:idx] +
                '<strong>' + nombre[idx:idx+len(term)] + '</strong>' +
                nombre[idx+len(term):]
            )
        else:
            highlighted = nombre
            
        results.append({
            'id': enfermedad.id_enfermedad,
            'text': nombre,
            'highlighted': highlighted
        })
    
    return JsonResponse(results, safe=False)

def buscar_pacientes2(request):
    query = request.GET.get('q', '')
    resultados = []

    if query:
        pacientes = Paciente.objects.filter(
            Q(user__first_name__icontains=query) | 
            Q(user__last_name__icontains=query) |
            Q(user__username__icontains=query)
        ).select_related('user')[:10]  # Limita a 10 resultadosAdd commentMore actions

        resultados = [
            {
                'id': paciente.id,
                'nombre': f"{paciente.user.first_name} {paciente.user.last_name}",
                'username': paciente.user.username  # Añadimos el username
            }
            for paciente in pacientes
        ]

    return JsonResponse(resultados, safe=False)

def datos_tendencias_enfermedades(request):
    try:
        enfermedad_id = request.GET.get('enfermedad_id')
        paciente_username = request.GET.get('paciente_username')
        
        response_data = {
            "success": True,
            "title": "Tendencias de Enfermedades",
            "labels": [],
            "values": [],
            "periodo": "Año actual"
        }
        
        # Obtener el rango del año actual (desde enero hasta hoy)
        today = timezone.now().date()
        start_date = date(today.year, 1, 1)  # 1 de enero del año actual
        end_date = today
        
        # Datos para la enfermedad seleccionada
        if enfermedad_id:
            try:
                enfermedad = Enfermedad.objects.get(id_enfermedad=enfermedad_id)
                
                # Consulta corregida para total_pacientes
                total_pacientes = Registro.objects.filter(
                    id_enfermedad=enfermedad,
                    id_cita__fecha__range=(start_date, end_date)
                ).values('id_cita__id_paciente').distinct().count()
                
                response_data['enfermedad_info'] = {
                    'id': enfermedad.id_enfermedad,
                    'nombre': enfermedad.nombre,
                    'total_pacientes': total_pacientes
                }
                
                # Datos mensuales para la enfermedad
                registros = Registro.objects.filter(
                    id_enfermedad=enfermedad,
                    id_cita__fecha__range=(start_date, end_date)
                ).annotate(
                    month=ExtractMonth('id_cita__fecha'),
                    year=ExtractYear('id_cita__fecha')
                ).values('month', 'year').annotate(
                    count=Count('id_registro')
                ).order_by('year', 'month')
                
                # Crear labels con nombres de meses en español
                meses_espanol = [
                    'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                    'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'
                ]
                
                # Inicializar todos los meses del año con 0
                datos_mensuales = {mes: 0 for mes in range(1, 13)}
                
                # Actualizar con los datos reales
                for r in registros:
                    datos_mensuales[r['month']] = r['count']
                
                # Preparar datos para el gráfico
                meses = []
                conteos = []
                
                for mes_num in range(1, 13):
                    meses.append(meses_espanol[mes_num - 1])
                    conteos.append(datos_mensuales[mes_num])
                    # Solo incluir meses hasta el actual si estamos en el año actual
                    if mes_num == today.month and today.year == timezone.now().year:
                        break
                
                response_data['labels'] = meses
                response_data['values'] = conteos
                response_data['title'] = f"Casos de {enfermedad.nombre} en {today.year}"
                
            except Enfermedad.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "Enfermedad no encontrada"
                }, status=404)
        
        # Resto del código para paciente seleccionado...
        if paciente_username:
            try:
                paciente = Paciente.objects.get(user__username=paciente_username)
                
                # Consulta corregida para total_registros
                total_registros = Registro.objects.filter(
                    id_cita__id_paciente=paciente,
                    id_cita__fecha__range=(start_date, end_date)
                ).count()
                
                response_data['paciente_info'] = {
                    'id': paciente.id,
                    'nombre': f"{paciente.user.first_name} {paciente.user.last_name}",
                    'total_registros': total_registros
                }
                
                # Enfermedades más comunes del paciente en el año actual
                if not enfermedad_id:
                    enfermedades = Registro.objects.filter(
                        id_cita__id_paciente=paciente,
                        id_cita__fecha__range=(start_date, end_date)
                    ).values('id_enfermedad__nombre')\
                     .annotate(count=Count('id_enfermedad'))\
                     .order_by('-count')[:5]
                    
                    labels = []
                    values = []
                    
                    for e in enfermedades:
                        labels.append(e['id_enfermedad__nombre'])
                        values.append(e['count'])
                    
                    response_data['labels'] = labels
                    response_data['values'] = values
                    response_data['title'] = f"Enfermedades más frecuentes de {paciente.user.first_name} en {today.year}"
                    
            except Paciente.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "Paciente no encontrado"
                }, status=404)
        
        # Validar que hay datos para mostrar
        if not response_data['labels'] and not response_data['values']:
            response_data.update({
                "success": False,
                "message": "No se encontraron datos para los filtros aplicados"
            })
        
        return JsonResponse(response_data)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            "success": False,
            "message": f"Error interno del servidor: {str(e)}"
        }, status=500)

    
def filtro_enfermedades(request):
    try:
        nombre_enfermedad = request.GET.get('nombre_enfermedad')
        id_enfermedad = request.GET.get('id_enfermedad')
        paciente_username = request.GET.get('paciente_username')

        print(f"Parámetros recibidos - nombre_enfermedad: {nombre_enfermedad}, id_enfermedad: {id_enfermedad}, paciente_username: {paciente_username}")

        # Obtener la enfermedad
        enfermedad = None
        if id_enfermedad:
            enfermedad = Enfermedad.objects.filter(id_enfermedad=id_enfermedad).first()
            print(f"Enfermedad encontrada por ID: {enfermedad}")
        elif nombre_enfermedad:
            enfermedad = Enfermedad.objects.filter(nombre__icontains=nombre_enfermedad).first()
            print(f"Enfermedad encontrada por nombre: {enfermedad}")

        # Datos base para la respuesta
        response_data = {
            "success": True,
            "enfermedad": {
                'id': enfermedad.id_enfermedad if enfermedad else None,
                'nombre': enfermedad.nombre if enfermedad else None
            } if enfermedad else None,
            "registros": [],
            "tiene_enfermedad": False
        }

        # Caso 1: Solo paciente (sin enfermedad)
        if paciente_username and not enfermedad:
            print("Caso 1: Solo paciente")
            try:
                paciente = Paciente.objects.get(user__username=paciente_username)
                print(f"Paciente encontrado: {paciente}")
                
                registros = Registro.objects.filter(
                    id_cita__id_paciente=paciente
                ).select_related('id_enfermedad', 'id_cita').prefetch_related(
                    'registrotratamiento_set__id_tratamiento',
                    'registroproducto_set__id_producto'
                )
                print(f"Total registros encontrados: {registros.count()}")
                
                registros_data = []
                for r in registros:
                    print(f"Procesando registro ID: {r.id_registro}")
                    tratamientos = [{
                        'id': rt.id_tratamiento.id_tratamiento,
                        'nombre': rt.id_tratamiento.nombre
                    } for rt in r.registrotratamiento_set.all()]
                    print(f"Tratamientos encontrados: {len(tratamientos)}")
                    
                    productos = [{
                        'id': rp.id_producto.id_producto,
                        'nombre': rp.id_producto.nombre
                    } for rp in r.registroproducto_set.all()]
                    print(f"Productos encontrados: {len(productos)}")
                    
                    registros_data.append({
                        'id': r.id_registro,
                        'motivo': r.motivo,
                        'fecha': r.id_cita.fecha.strftime('%Y-%m-%d') if r.id_cita.fecha else None,
                        'enfermedad': {
                            'id': r.id_enfermedad.id_enfermedad,
                            'nombre': r.id_enfermedad.nombre
                        },
                        'tratamientos': tratamientos,
                        'productos': productos,
                        'asistencia': r.id_cita.estado == 'finalizada'
                    })
                
                response_data.update({
                    "registros_paciente": registros.count(),
                    "registros": registros_data,
                    "paciente": {
                        "id": paciente.id,
                        "nombre": f"{paciente.user.first_name} {paciente.user.last_name}"
                    }
                })
                
            except Paciente.DoesNotExist as e:
                print(f"Error: Paciente no encontrado - {str(e)}")
                response_data.update({
                    "success": False,
                    "message": "Paciente no encontrado"
                })
            except Exception as e:
                print(f"Error inesperado en Caso 1: {str(e)}")
                raise e
                
            return JsonResponse(response_data)

        # Caso 2: Solo enfermedad (sin paciente)
        if enfermedad and not paciente_username:
            print("Caso 2: Solo enfermedad")
            try:
                total_registros = Registro.objects.filter(id_enfermedad=enfermedad).count()
                pacientes_unicos = Registro.objects.filter(
                    id_enfermedad=enfermedad
                ).values('id_cita__id_paciente').distinct().count()
                
                print(f"Total registros: {total_registros}, Pacientes únicos: {pacientes_unicos}")
                
                response_data.update({
                    "total_registros": total_registros,
                    "total_pacientes_unicos": pacientes_unicos
                })
            except Exception as e:
                print(f"Error inesperado en Caso 2: {str(e)}")
                raise e
                
            return JsonResponse(response_data)

        # Caso 3: Enfermedad Y paciente
        if enfermedad and paciente_username:
            print("Caso 3: Enfermedad y paciente")
            try:
                paciente = Paciente.objects.get(user__username=paciente_username)
                print(f"Paciente encontrado: {paciente}")
                
                registros_paciente = Registro.objects.filter(
                    id_cita__id_paciente=paciente,
                    id_enfermedad=enfermedad
                ).select_related('id_cita', 'id_enfermedad').prefetch_related(
                    'tratamientos__id_tratamiento',  # Cambiado de registrotratamiento_set a tratamientos
                    'registroproducto_set__id_producto'
                )
                print(f"Total registros encontrados: {registros_paciente.count()}")
                
                tiene_enfermedad = PacienteEnfermedad.objects.filter(
                    id_paciente=paciente,
                    id_enfermedad=enfermedad
                ).exists()
                print(f"Tiene enfermedad directamente: {tiene_enfermedad}")
                
                registros_data = []
                for r in registros_paciente:
                    print(f"Procesando registro ID: {r.id_registro}")
                    tratamientos = [{
                        'id': rt.id_tratamiento.id_tratamiento,
                        'nombre': rt.id_tratamiento.nombre
                    } for rt in r.tratamientos.all()]  # Cambiado de registrotratamiento_set a tratamientos
                    
                    productos = [{
                        'id': rp.id_producto.id_producto,
                        'nombre': rp.id_producto.nombre
                    } for rp in r.registroproducto_set.all()]
                    
                    registros_data.append({
                        'id': r.id_registro,
                        'motivo': r.motivo,
                        'fecha': r.id_cita.fecha.strftime('%Y-%m-%d') if r.id_cita.fecha else None,
                        'enfermedad': {
                            'id': r.id_enfermedad.id_enfermedad,
                            'nombre': r.id_enfermedad.nombre
                        },
                        'tratamientos': tratamientos,
                        'productos': productos,
                        'asistencia': r.id_cita.estado == 'finalizada'
                    })
                
                response_data.update({
                    "registros_paciente": registros_paciente.count(),
                    "tiene_enfermedad": tiene_enfermedad,
                    "registros": registros_data,
                    "paciente": {
                        "id": paciente.id,
                        "nombre": f"{paciente.user.first_name} {paciente.user.last_name}"
                    }
                })
                
            except Paciente.DoesNotExist as e:
                print(f"Error: Paciente no encontrado - {str(e)}")
                response_data.update({
                    "success": False,
                    "message": "Paciente no encontrado"
                })
            except Exception as e:
                print(f"Error inesperado en Caso 3: {str(e)}")
                raise e
                
            return JsonResponse(response_data)
        # Si no se cumplen los casos anteriores, retornamos un error
    except Exception as e:
        print(f"Error inesperado: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)
    
def filtrar_registros_tratamientos(request):
    try:
        nombre_enfermedad = request.GET.get('nombre_enfermedad', None)
        paciente_username = request.GET.get('paciente_username', None)
        nombre_tratamiento = request.GET.get('nombre_tratamiento', None)

        # validar parametros minimos
        if not nombre_enfermedad or not nombre_tratamiento:
            return JsonResponse({
                "success": False,
                "message": "Se requieren los parámetros 'enfermedad' y 'tratamiento'"
            }, status=400)
        
        # obtenemos la enferemdad y tratamiento
        enfermedad = Enfermedad.objects.filter(nombre__icontains=nombre_enfermedad).first()
        tratamiento = Tratamiento.objects.filter(nombre__icontains=nombre_tratamiento).first()

        if not enfermedad or not tratamiento:
            return JsonResponse({
                "success": False,
                "message": "Enfermedad o tratamiento no encontrado"
            }, status=404)
        
        # aqui construimos la consulta base
        registros = Registro.objects.filter(
            id_enfermedad=enfermedad,
            tratamientos__id_tratamiento=tratamiento
        ).select_related(
            'id_cita__id_paciente',
            'id_enfermedad',
        ).prefetch_related('tratamientos__id_tratamiento')

        # filtramos por paciente solo si se especifica
        if paciente_username:
            try:
                paciente = Paciente.objects.get(user__username=paciente_username)
                registros = registros.filter(id_cita__id_paciente=paciente)
            except Paciente.DoesNotExist:
                return JsonResponse({
                    "success": False,
                    "message": "Paciente no encontrado"
                }, status=404)
        
        # obteniendo los datos de los registros
        registros_data = []
        pacientes_unicos = set()  # Para evitar duplicados

        for registro in registros:
            # datos del paciente
            paciente = registro.id_cita.id_paciente
            pacientes_unicos.add(paciente.id)

            # datos del registro
            registros_data.append({
                'id': registro.id_registro,
                'fecha': registro.id_cita.fecha,
                'motivo': registro.motivo,
                'paciente': {
                    'id': paciente.id,
                    'nombre_completo': f"{paciente.user.first_name} {paciente.user.last_name}",
                    'username': paciente.user.username,
                },
                'tratamientos': [
                    {
                        'id': rt.id_tratamiento.id_tratamiento,  # Accedemos al id del tratamiento
                        'nombre': rt.id_tratamiento.nombre,      # Accedemos al nombre del tratamiento
                    } for rt in registro.tratamientos.all()
                ]
            })

        # preparamos la respuesta JSON
        response_data = {
            "success": True,
            "enfermedad": {
                'id': enfermedad.id_enfermedad,
                'nombre': enfermedad.nombre
            },
            "tratamiento": {
                "id": tratamiento.id_tratamiento,
                "nombre": tratamiento.nombre
            },
            "total_registros": len(registros_data),  # Usamos len() en lugar de count() para usar los datos ya obtenidos
            "total_pacientes_unicos": len(pacientes_unicos),
            "registros": registros_data,
        }

        # si se filtro por paciente, agregamos info adicional
        if paciente_username:
            response_data['paciente'] = registros_data[0]['paciente'] if registros_data else None
            response_data['tiene_enfermedad'] = PacienteEnfermedad.objects.filter(
                id_paciente=paciente,
                id_enfermedad=enfermedad
            ).exists()

        return JsonResponse(response_data)

    except Exception as e:
        print(f"Error en filtrar_registros_tratamientos: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

def estadisticas_tratamientos(request):
    try:
        # Obtenemos todos los tratamientos con sus estadísticas
        tratamientos = Tratamiento.objects.annotate(
            total_registros=Count('registrotratamiento'),
            total_pacientes=Count('registrotratamiento__id_registro__id_cita__id_paciente', distinct=True),
            total_enfermedades=Count('tratamientoenfermedad__id_enfermedad', distinct=True)
        ).prefetch_related(
            'tratamientoenfermedad_set__id_enfermedad'
        )

        tratamientos_data = []
        for tratamiento in tratamientos:
            # Obtenemos las enfermedades relacionadas
            enfermedades = [{
                'id': te.id_enfermedad.id_enfermedad,
                'nombre': te.id_enfermedad.nombre
            } for te in tratamiento.tratamientoenfermedad_set.all()]
            
            tratamientos_data.append({
                'id': tratamiento.id_tratamiento,
                'nombre': tratamiento.nombre,
                'descripcion': tratamiento.descripcion,
                'total_registros': tratamiento.total_registros,
                'total_pacientes': tratamiento.total_pacientes,
                'enfermedades': enfermedades,
                'enfermedades_lista': ", ".join([e['nombre'] for e in enfermedades])
            })

        return JsonResponse({
            "success": True,
            "tratamientos": tratamientos_data,
            "total_tratamientos": len(tratamientos_data)
        })

    except Exception as e:
        print(f"Error en estadisticas_tratamientos: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)
    
def enfermedades_por_tratamiento(request):
    tratamiento_id = request.GET.get('tratamiento_id')
    
    if not tratamiento_id:
        return JsonResponse({
            "success": False,
            "message": "Se requiere el parámetro tratamiento_id"
        }, status=400)
    
    try:
        # Obtenemos las enfermedades relacionadas con el tratamiento
        relaciones = TratamientoEnfermedad.objects.filter(
            id_tratamiento=tratamiento_id
        ).select_related('id_enfermedad')
        
        enfermedades = [{
            'id': rel.id_enfermedad.id_enfermedad,
            'nombre': rel.id_enfermedad.nombre
        } for rel in relaciones]
        
        return JsonResponse({
            "success": True,
            "enfermedades": enfermedades
        })
        
    except Exception as e:
        print(f"Error en enfermedades_por_tratamiento: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

def tendencias_tratamientos(request):
    tratamiento_id = request.GET.get('tratamiento_id')
    enfermedad_id = request.GET.get('enfermedad_id')
    
    try:
        # Consulta base usando el ORM de Django
        queryset = RegistroTratamiento.objects.select_related('id_tratamiento')
        
        # Aplicar filtros según los parámetros
        if tratamiento_id:
            queryset = queryset.filter(id_tratamiento_id=tratamiento_id)
            
        if enfermedad_id:
            queryset = queryset.filter(
                id_tratamiento__tratamientoenfermedad__id_enfermedad_id=enfermedad_id
            )
        
        # Anotar con el mes y contar las ocurrencias
        datos_tendencia = queryset.annotate(
            mes=TruncMonth('id_registro__id_cita__fecha')
        ).values('mes', 'id_tratamiento__nombre').annotate(
            cantidad=Count('id')
        ).order_by('mes')
        
        # Formatear los datos para la respuesta
        datos = [{
            'mes': item['mes'].strftime('%Y-%m'),
            'cantidad': item['cantidad'],
            'tratamiento_nombre': item['id_tratamiento__nombre']
        } for item in datos_tendencia]
        
        return JsonResponse({
            "success": True,
            "datos": datos,
            "filtros": {
                "tratamiento_id": tratamiento_id,
                "enfermedad_id": enfermedad_id
            }
        })
        
    except Exception as e:
        print(f"Error en tendencias_tratamientos: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": "Error interno del servidor"
        }, status=500)

# endpoint que te devuelve todas las citas, esto es inicialmente para el inicio de visualizacion
def citas_visualizacion(request):
    citas = Cita.objects.all().order_by('-fecha', '-hora').select_related(
        'id_paciente__user',
        'id_medico__user',
        'id_especialidad'
    )
    
    citas_data = []
    for cita in citas:
        citas_data.append({
            "id_cita": cita.id_cita,
            "id_paciente_id": cita.id_paciente.id,
            "paciente_nombre": f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
            "id_medico_id": cita.id_medico.id,
            "medico_nombre": f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
            "fecha": cita.fecha,
            "hora": cita.hora.strftime("%H:%M:%S"),
            "estado": cita.estado
        })
    
    return JsonResponse({
        "citas": citas_data,
        "success": True,
        "count": len(citas_data)
    }, encoder=DjangoJSONEncoder)

# ENDPOINT QUE ME DEVUELVE A TODOS LOS DOCTORES
def doctores_all(request):
    doctores = Medico.objects.all().select_related('user')
    doctores_data = []
    for doctor in doctores:
        doctores_data.append({
            "id": doctor.id,
            "nombre": f"{doctor.user.first_name} {doctor.user.last_name}"
        })
    
    return JsonResponse({
        "doctores": doctores_data,
        "success": True,
        "count": len(doctores_data)
    }, encoder=DjangoJSONEncoder)

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from .models import Cita, Paciente, Medico, Especialidad
import json

from django.core.paginator import Paginator

def tipo_citas_all(request, tipo):
    citas = []

    citas = Cita.objects.all().order_by('-fecha', '-hora')
    #print(f"Cantidad de citas: {citas.count()}")
    
    if tipo == 'pendientes':
        queryset = Cita.objects.filter(estado='pendiente')
    elif tipo == 'finalizadas':
        queryset = Cita.objects.filter(estado='finalizada')
    elif tipo == 'no_asistio':
        queryset = Cita.objects.filter(estado='no_asistio')
    else:
        return JsonResponse({"success": False, "message": "Tipo no válido"}, status=400)
    
    # Paginación
    page_number = request.GET.get('page', 1)
    paginator = Paginator(queryset.select_related('id_paciente__user', 'id_medico__user'), 10)
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Serializar los datos
    citas_list = list(page_obj.object_list.values(
        'id_cita', 'fecha', 'hora', 'estado',
        'id_paciente__user__first_name', 'id_paciente__user__last_name',
        'id_paciente__telefono',
        'id_medico__user__first_name', 'id_medico__user__last_name'
    ))
    
    return JsonResponse({
        "success": True,
        "citas": citas_list,
        "count": paginator.count,
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "current_page": page_obj.number,
        "total_pages": paginator.num_pages
    })


#@login_required
@require_GET
def todas_las_citas_json(request):
    # Obtener todas las citas con sus relaciones
    citas = Cita.objects.select_related(
        'id_paciente__user',
        'id_medico__user',
        'id_especialidad'
    ).all()
    
    # Preparar los datos para el JSON
    citas_data = []
    for cita in citas:
        citas_data.append({
            'id_cita': cita.id_cita,
            'paciente': {
                'id': cita.id_paciente.id,
                'nombre': f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
                'cedula': cita.id_paciente.cedula,
                'telefono': cita.id_paciente.telefono,
            },
            'medico': {
                'id': cita.id_medico.id,
                'nombre': f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
            },
            'especialidad': {
                'id': cita.id_especialidad.id_especialidad,
                'nombre': cita.id_especialidad.nombre,
            },
            'fecha': cita.fecha.strftime('%Y-%m-%d'),
            'hora': cita.hora.strftime('%H:%M:%S'),
            'estado': cita.estado,
            'estado_display': cita.get_estado_display(),
        })
    
    # Devolver la respuesta JSON
    return JsonResponse({
        'success': True,
        'count': len(citas_data),
        'citas': citas_data,
    }, safe=False, json_dumps_params={'ensure_ascii': False, 'indent': 2})

def generar_reporte(request):

    citas = Cita.objects.select_related('id_paciente__user', 'id_medico__user', 'id_especialidad').all()

    # Paleta de colores verde
    # Paleta de colores verde
    COLOR_ENCABEZADO = '006400'  # Verde oscuro
    COLOR_FONDO = 'E8F5E9'       # Verde claro
    COLOR_BORDE = '2E7D32'       # Verde medio

    # Estilos definidos
    estilo_encabezado = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    estilo_normal = Font(name='Calibri', size=11)
    relleno_encabezado = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type='solid')
    relleno_fila = PatternFill(start_color=COLOR_FONDO, end_color=COLOR_FONDO, fill_type='solid')
    borde_fino = Side(border_style='thin', color=COLOR_BORDE)
    borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    alineacion_centro = Alignment(horizontal='center', vertical='center')

    for i in citas:
        print(f"Paciente: {i.id_paciente.user.first_name} -- Medico: {i.id_medico.user.first_name} -- fecha: {i.fecha} -- hora: {i.hora} -- estado: {i.estado}")
    
    encabezados_citas = ['Cedula' ,'Id Cita' ,'Nombre Paciente', 'Apellido Paciente', 'Estado', 'Fecha', 'Hora', 'Medico', 'Medico_user' ,'Especialidad']
    years = []
    for i in citas:
        if i.fecha:  # Asegúrate de que la fecha no sea None
            year = i.fecha.year  # Lee el año
            if year not in years:
                years.append(year)
    print(f"Years: {years}")

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar la hoja activa por defecto
    

    # iteramos en cada year, cada year es una hoja de excel
    for year in years:
        hoja_citas = wb.create_sheet(title=str(year))

        hoja_citas.append(encabezados_citas)

        # Aplicando estilos a encabezados
        for col in range(1, len(encabezados_citas) + 1):
            celda = hoja_citas.cell(row=1, column=col)
            celda.font = estilo_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_centro
            celda.border = borde_celda

        # agregar datos
        fila_num = 2

        for cita in citas:
            if cita.fecha.year == year:
                hoja_citas.append([
                    cita.id_paciente.cedula,
                    cita.id_cita,
                    cita.id_paciente.user.first_name,
                    cita.id_paciente.user.last_name,
                    cita.estado,
                    cita.fecha.strftime('%Y-%m-%d'),
                    cita.hora.strftime('%H:%M:%S'),
                    cita.id_medico.user.first_name,
                    cita.id_medico.user.username,
                    cita.id_especialidad.nombre
                ])

                # aplicar datos a la fila de datos
                for col in range(1, len(encabezados_citas) + 1):
                    celda = hoja_citas.cell(row=fila_num, column=col)
                    celda.font = estilo_normal
                    celda.alignment = alineacion_centro
                    celda.border = borde_celda

                    # Alternar colores de fila
                    if fila_num % 2 == 0:
                        celda.fill = relleno_fila
                
                fila_num += 1

        # ajustando ancho de las columnas
        for col in range(1, len(encabezados_citas) + 1):
            col_letter = get_column_letter(col)
            hoja_citas.column_dimensions[col_letter].width = max(
                len(encabezados_citas[col-1]) + 2,  # Ancho mínimo basado en encabezado
                15  # Ancho mínimo predeterminado
            )
        # Congelar encabezados para scroll
        hoja_citas.freeze_panes = 'A2'

    # preparar la respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_citas.xlsx"'

    wb.save(response)

    return response


# Historial paciente - todas las citas del paciente
#@login_required
def historial_paciente(request, username):
    # Obtener el paciente por username
    paciente = get_object_or_404(Paciente, user__username=username)
    citas = Cita.objects.filter(id_paciente=paciente).order_by('-fecha', '-hora')
    
    cita_data = []
    for cita in citas:
        cita_data.append({
            'id_cita': cita.id_cita,
            'fecha': cita.fecha.strftime('%Y-%m-%d'),
            'hora': cita.hora.strftime('%H:%M:%S'),
            'estado': cita.estado,
            'id_medico': cita.id_medico.id,
            'nombre_medico': f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}",
            'especialidad': cita.id_especialidad.nombre,
            'motivo': getattr(cita, 'motivo', 'Consulta médica'),  # Forma más segura de getattr
            'notas': getattr(cita, 'notas', 'Sin notas adicionales'),
            'tipo': getattr(cita, 'tipo', 'Consulta')  # Agregado tipo por si lo necesitas

        })

    for cita in citas:
        print(f"Paciente-user:  {cita.id_paciente.user.username} -- Paciente: {cita.id_paciente.user.first_name } {cita.id_paciente.user.last_name}  -- Medico: {cita.id_medico.user.first_name} -- fecha: {cita.fecha} -- hora: {cita.hora} -- estado: {cita.estado}")
        print("--------------------------------------------------")
    
    if citas:
        return JsonResponse({"success": True, "data": cita_data})
    
    return JsonResponse({"success": False, "message": "No se encontraron citas para este paciente."})

# vista personalizada para el logout
def logout_user(request):

    #return HttpResponse(f"{request.user.role}")
    if hasattr(request.user, 'role'):
        user_role = request.user.role
    else:
        user_role = None
    
    logout(request)

    return redirect('login')

# vista del dashboard del medico
@login_required(login_url='login_medico')
def dashboard_medico(request):
    if hasattr(request.user, 'role') and request.user.role == 'Medico':
        return render(request, 'citas/dashboard_medico.html')
    else:
        return redirect('login_medico')
    

# vista del dashboard del paciente
@login_required(login_url='login_paciente')
def dashboard_paciente(request):
    if  hasattr(request.user, 'role') and request.user.role == 'Paciente':
        return render(request, 'citas/dashboard_paciente.html')
    else:
        return redirect('login_paciente')

# vista para crear alergias
def crear_alergia(request):
    if request.method == "POST":
        # datos de la alergia
        nombre = request.POST['nombre']
        descripcion = request.POST['descripcion']

        print(f"Nombre: {nombre}")
        print(f"Descripcion: {descripcion}")
        #return HttpResponse("Alergia creada")

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('crear_alergia')

        if Alergia.objects.filter(nombre=nombre).exists():
            messages.error(request, 'La alergia ya existe.')
            return redirect('crear_alergia')

        alergia = Alergia.objects.create(
            nombre=nombre,
            descripcion=descripcion
        )

        messages.success(request, 'La alergia se ha creado correctamente.')
        return redirect('enfermedades_alergias')
    #return render(request, 'citas/crear_alergia.html')

@login_required
def listar_alergias(request):
    alergias = Alergia.objects.all()
    return render(request, 'citas2/alergias.html', {'alergias': alergias})

# crear especialidades
def crear_especialidad(request):
    if request.method == "POST":
        # datos de la especialidad
        nombre = request.POST['nombre']

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('crear_especialidad')

        if Especialidad.objects.filter(nombre=nombre).exists():
            messages.error(request, 'La especialidad ya existe.')
            return redirect('crear_especialidad')

        Especialidad.objects.create(
            nombre=nombre,
        )

        messages.success(request, 'La especialidad se ha creado correctamente.')
        return redirect('crear_especialidad')
    
    return render(request, 'citas/crear_especialidad.html')

def vista_error(request):
    return render(request, 'citas2/vista_error.html')

#@admin_medico_required
@login_required
def listar_medicos(request):
    
    medicos = Medico.objects.prefetch_related('medicoespecialidad_set__id_especialidad')
    especialidades = Especialidad.objects.all()
    print(f"Medicos: {medicos}")
    return render(request, 'citas2/medicos.html', {'medicos': medicos, 'especialidades': especialidades})

def detalle_medico(request, medico_id):
    try:
        medico = Medico.objects.get(id=medico_id)
        especialidades = MedicoEspecialidad.objects.filter(id_medico=medico)
    except Medico.DoesNotExist:
        messages.error(request, 'El médico no existe.')
        return redirect('listar_medicos')

    return render(request, 'citas/detalle_medico.html', {'medico': medico, 'especialidades': especialidades})

def asignar_especialidad(request, medico_id):
    try:
        medico = Medico.objects.get(id=medico_id)
    except Medico.DoesNotExist:
        messages.error(request, 'El medico no existe.')
        return redirect('listar_medicos')
    
    especialidades = Especialidad.objects.all()

    if request.method == "POST":
        especialidades_seleccionadas = request.POST.getlist('especialidades')

        for especialidad_id in especialidades_seleccionadas:
            especialidad = Especialidad.objects.get(id_especialidad=especialidad_id)
            if not MedicoEspecialidad.objects.filter(id_medico=medico, id_especialidad=especialidad).exists():
                MedicoEspecialidad.objects.create(id_medico=medico, id_especialidad=especialidad)
            else:
                messages.warning(
                    request, f"La especialidad '{especialidad.nombre}' ya está asignada al médico."
                )
                return redirect('asignar_especialidad', medico_id=medico_id)
        
        messages.success(request, 'Las especialidades se han asignado correctamente.')
        return redirect('detalle_medico', medico_id=medico_id)
    
    return render(request, 'citas/asignar_especialidad.html', {'medico': medico, 'especialidades': especialidades})

# vistas para enfermedades
@admin_medico_required
@login_required
def listar_enfermedades_alergias(request):
    enfermedades = Enfermedad.objects.all()
    alergias = Alergia.objects.all()
    # ruta_original = citas/listar_enfermedades.html
    return render(request, 'citas2/enfermedades-alergia.html', {'enfermedades': enfermedades, 'alergias': alergias})

def crear_enfermedad(request):
    if request.method == "POST":
        # datos de la enfermedad
        nombre = request.POST['nombre']
        descripcion = request.POST['descripcion']

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('crear_enfermedad')
        
        print(f"Nombre: {nombre}")
        print(f"Descripcion: {descripcion}")
        #return HttpResponse("Enfermedad agregada")

        if Enfermedad.objects.filter(nombre=nombre).exists():
            messages.error(request, 'La enfermedad ya existe.')
            return redirect('crear_enfermedad')
        
        Enfermedad.objects.create(
            nombre=nombre,
            descripcion=descripcion
        )

        messages.success(request, 'La enfermedad se ha creado correctamente.')
        return redirect('enfermedades_alergias')
    
    return render(request, 'citas/crear_enfermedad.html')

# vistas para tratamentos
@login_required
def listar_tratamientos(request):
    tratamientos = Tratamiento.objects.prefetch_related('tratamientoenfermedad_set__id_enfermedad')
    print(f"Tratamientos: {tratamientos}")

    # mostrando la enfermedad que trata cada tratamiento
    for tratamiento in tratamientos:
        print(f"Tratamiento: {tratamiento} - Enfermedades: {tratamiento.tratamientoenfermedad_set.all()}")
    return render(request, 'citas2/tratamientos.html', {'tratamientos': tratamientos})

# buscar enfermedades
@login_required
def buscar_enfermedades(request):
    if request.method == 'GET':
        query = request.GET.get("q", "")
        if query:
            enfermedades = Enfermedad.objects.filter(
                nombre__icontains=query
            ).values('id_enfermedad', 'nombre')

            return JsonResponse(list(enfermedades), safe=False)
        
        return JsonResponse([], safe=False)


@login_required
def crear_tratamiento(request):
    if request.method == "POST":
        # datos del tratamiento
        nombre = request.POST['nombre']
        descripcion = request.POST['descripcion']
        id_enfermedad = request.POST.get('id_enfermedad')

        print(f"Nombre: {nombre}")
        print(f"Descripcion: {descripcion}")
        print(f"id_enfermedad: {id_enfermedad}")

        # verificamos si la enfermedad existe
        enfermedad = get_object_or_404(Enfermedad, id_enfermedad=id_enfermedad)

        #return HttpResponse("tratamiento creado")

        # verificar si ya existe un tratamiento con el mismo nombre
        if Tratamiento.objects.filter(nombre=nombre).exists():
            messages.error(request, 'El tratamiento ya existe.')
            return redirect('tratamientos')
        

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
            return redirect('tratamientos')

        if Tratamiento.objects.filter(nombre=nombre).exists():
            messages.error(request, 'El tratamiento ya existe.')
            return redirect('tratamientos')
        
        # creando el tratamiento
        tratamiento = Tratamiento.objects.create(
            nombre=nombre,
            descripcion=descripcion
        )

        # asociar el tratamiento con la enfermedad a tratar
        TratamientoEnfermedad.objects.create(
            id_tratamiento=tratamiento,
            id_enfermedad=enfermedad
        )

        messages.success(request, 'El tratamiento se ha creado correctamente.')
        return redirect('tratamientos')
    
    #return render(request, 'citas/crear_tratamiento.html')

@login_required
def asignar_tratamiento(request):
    if request.method == "POST":
        tratamiento_id = request.POST.get('tratamiento')
        enfermedad_id = request.POST.get('enfermedad')

        #tratamiento = Tratamiento.objects.get(id_tratamiento=tratamiento_id)
        if not enfermedad_id or not tratamiento_id:
            messages.error(request, 'Debe seleccionar una enfermedad y un tratamiento.')
            return redirect('asignar_tratamiento')
        
        try:
            tratamiento = Tratamiento.objects.get(id_tratamiento=tratamiento_id)
            enfermedad = Enfermedad.objects.get(id_enfermedad=enfermedad_id)
        except (Tratamiento.DoesNotExist, Enfermedad.DoesNotExist):
            messages.error(request, 'El tratamiento o enfermedad no existe.')
            return redirect('asignar_tratamiento')

        if TratamientoEnfermedad.objects.filter(id_tratamiento=tratamiento, id_enfermedad=enfermedad).exists():
            messages.error(request, 'El tratamiento ya está asignado a la enfermedad.')
            return redirect('asignar_tratamiento')

        TratamientoEnfermedad.objects.create(
            id_tratamiento=tratamiento,
            id_enfermedad=enfermedad
        )

        messages.success(request, 'El tratamiento se ha asignado correctamente.')
        return redirect('asignar_tratamiento')
    
    tratamientos = Tratamiento.objects.all()
    enfermedades = Enfermedad.objects.all()
    return render(request, 'citas/asignar_tratamiento.html', {'tratamientos': tratamientos, 'enfermedades': enfermedades})

def listar_enfermedades_tratamientos(request):
    enfermedades = Enfermedad.objects.all()
    tratamientos_asociados = {
        enfermedad: TratamientoEnfermedad.objects.filter(id_enfermedad=enfermedad) for enfermedad in enfermedades
    }

    return render(request, 'citas/listar_enfermedades_tratamientos.html', {'enfermedades': enfermedades, 'tratamientos_asociados': tratamientos_asociados})

@login_required
def agregar_producto(request):
    tipos = ["Cremas", "Geles", "Lociones", "Sueros", "Protector solar", "Jabón dermatológico"]

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        precio = request.POST.get('precio')
        marca = request.POST.get('marca')
        tipo = request.POST.get('tipo')

        print(f"Nombre: {nombre}")
        print(f"Descripción: {descripcion}")
        print(f"Precio: {precio}")
        print(f"Marca: {marca}")
        print(f"Tipo: {tipo}")

        #return HttpResponse("producto agregado")

        # Validaciones básicas
        if not nombre or not precio or not marca or not tipo:
            messages.error(request, "Por favor, completa todos los campos obligatorios.")
        else:
            try:
                # Crear el producto
                producto = Producto(
                    nombre=nombre,
                    descripcion=descripcion,
                    precio=precio,
                    marca=marca,
                    tipo=tipo
                )
                producto.save()
                messages.success(request, "Producto agregado exitosamente.")
                return redirect('productos')  # Redirigir a la lista de productos
            except Exception as e:
                messages.error(request, f"Ocurrió un error al guardar el producto: {e}")

    #return render(request, 'citas/agregar_producto.html', {'tipos': tipos})

@admin_medico_required
@login_required
def listar_productos(request):
    productos = Producto.objects.all()
    tipos = ["Cremas", "Geles", "Lociones", "Sueros", "Protector solar", "Jabón dermatológico"]
    print(f"Productos: {productos}")
    return render(request, 'citas2/productos.html', {'productos': productos, 'tipos': tipos})

# editar citas
def editar_cita(request):
    if request.method == "POST":
        id_cita = request.POST.get('id_cita')
        nuevo_estado = request.POST.get('estado')

        print(f"Id_cita: {id_cita}")
        print(f"Nuevo estado: {nuevo_estado}")

        #return HttpResponse("Cita editada correctamente")

        # validar datos
        if not id_cita or not nuevo_estado:
            messages.error(request, "Por favor, completa todos los campos obligatorios.")
            return redirect('editar_cita')
        
        try:
            cita = Cita.objects.get(id_cita=id_cita)
        except Cita.DoesNotExist:
            messages.error(request, "La cita no existe.")
            return redirect('inicio')
        
        cita.estado = nuevo_estado
        cita.save()
        messages.success(request, "Cita actualizada exitosamente.")
        return redirect('inicio')

# crear citas
@login_required
def crear_cita(request):
    if request.method == "POST":
        id_paciente = request.POST.get('id_paciente')
        id_medico = request.POST.get('id_medico')
        id_especialidad = request.POST.get('id_especialidad')
        fecha_str = request.POST.get('fecha')
        hora_str = request.POST.get('hora_inicio')

        # Validaciones básicas
        if not all([id_paciente, id_medico, id_especialidad, fecha_str, hora_str]):
            messages.warning(request, "Por favor, completa todos los campos obligatorios.")
            return redirect('inicio')
        
        try:
            # Convertir strings a objetos de fecha/hora
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            hora_inicio = datetime.strptime(hora_str, '%H:%M').time()
            
            # Validar que la especialidad del médico esté asociada al médico
            if not MedicoEspecialidad.objects.filter(
                id_medico=id_medico, 
                id_especialidad=id_especialidad
            ).exists():
                messages.error(request, "La especialidad seleccionada no está asociada al médico.")
                return redirect('inicio')

            # Calcular hora de fin (asumiendo 1 hora de duración)
            hora_fin = (datetime.combine(fecha, hora_inicio) + timedelta(hours=1)).time()

            # 1. Verificar colisiones con citas existentes del MÉDICO
            citas_medico = Cita.objects.filter(
                id_medico=id_medico,
                fecha=fecha,
                estado__in=['pendiente', 'finalizada']
            ).exclude(estado='eliminada')

            for cita in citas_medico:
                cita_hora_fin = (datetime.combine(fecha, cita.hora) + timedelta(hours=1)).time()
                
                if (hora_inicio < cita_hora_fin) and (hora_fin > cita.hora):
                    messages.warning(
                        request, 
                        f"El médico ya tiene una cita programada de {cita.hora.strftime('%H:%M')} a {cita_hora_fin.strftime('%H:%M')}."
                    )
                    return redirect('inicio')

            # 2. Verificar colisiones con citas existentes del PACIENTE
            citas_paciente = Cita.objects.filter(
                id_paciente=id_paciente,
                fecha=fecha,
                estado__in=['pendiente', 'finalizada']
            ).exclude(estado='eliminada')

            for cita in citas_paciente:
                cita_hora_fin = (datetime.combine(fecha, cita.hora) + timedelta(hours=1)).time()
                
                if (hora_inicio < cita_hora_fin) and (hora_fin > cita.hora):
                    messages.warning(
                        request, 
                        f"El paciente ya tiene una cita programada de {cita.hora.strftime('%H:%M')} a {cita_hora_fin.strftime('%H:%M')} Dr. {cita.id_medico.user.get_full_name()}."
                    )
                    return redirect('inicio')

            # Si no hay colisiones, crear la cita
            paciente = get_object_or_404(Paciente, pk=id_paciente)
            medico = get_object_or_404(Medico, pk=id_medico)
            especialidad = get_object_or_404(Especialidad, pk=id_especialidad)
            
            Cita.objects.create(
                id_paciente=paciente,
                id_medico=medico,
                id_especialidad=especialidad,
                fecha=fecha,
                hora=hora_inicio
            )

            messages.success(request, "Cita creada exitosamente.")
            return redirect('inicio')

        except ValueError as e:
            messages.error(request, f"Error en el formato de fecha u hora: {str(e)}")
            return redirect('inicio')


def validar_disponibilidad_medico(request):
    id_medico = request.GET.get('id_medico')
    fecha_str = request.GET.get('fecha')
    hora_str = request.GET.get('hora')

    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hora_inicio = datetime.strptime(hora_str, '%H:%M').time()
        hora_fin = (datetime.combine(fecha, hora_inicio) + timedelta(hours=1)).time()

        citas = Cita.objects.filter(
            id_medico=id_medico,
            fecha=fecha,
            estado__in=['pendiente', 'finalizada']
        )

        for cita in citas:
            cita_hora_fin = (datetime.combine(fecha, cita.hora) + timedelta(hours=1)).time()
            if (hora_inicio < cita_hora_fin) and (hora_fin > cita.hora):
                return JsonResponse({
                    'disponible': False,
                    'mensaje': f"Ya existe una cita de {cita.hora.strftime('%H:%M')} a {cita_hora_fin.strftime('%H:%M')}."
                })

        return JsonResponse({'disponible': True})
    
    except Exception as e:
        return JsonResponse({'disponible': False, 'mensaje': 'Datos inválidos'})

def listar_citas(request):
    citas = Cita.objects.all()
    return render(request, 'citas/listar_citas.html', {'citas': citas})

# buscar por nombres de pacientes y medicos
@login_required
def buscar_pacientes(request):
    query = request.GET.get('q', '')
    pacientes = Paciente.objects.filter(user__first_name__icontains=query)[:10]
    resultados = [
        {'id': paciente.id, 'nombre': paciente.user.first_name, 'apellido': paciente.user.last_name}
        for paciente in pacientes
    ]

    print(f"Resultados: {resultados}")

    return JsonResponse(resultados, safe=False)

@login_required
def buscar_medicos(request):
    query = request.GET.get('q', '')
    medicos = Medico.objects.filter(user__first_name__icontains=query)[:10]
    resultados = [
        {'id': medico.id, 'nombre': medico.user.first_name, 'apellido': medico.user.last_name}
        for medico in medicos
    ]
    return JsonResponse(resultados, safe=False)

@login_required
def obtener_especialidades(request):
    medico_id = request.GET.get('id_medico')
    if medico_id:
        especialidades = MedicoEspecialidad.objects.filter(id_medico=medico_id).select_related('id_especialidad')
        especialidades_data = [{"id": e.id_especialidad.id_especialidad, "nombre": e.id_especialidad.nombre} for e in especialidades]
        return JsonResponse(especialidades_data, safe=False)
    return JsonResponse({"error": "Médico no encontrado"}, status=400)

@login_required
def buscar_cita(request):
    if request.method == "GET":
        codigo_cita = request.GET.get("codigo_cita", None)
        print(f"Codigo de cita: {codigo_cita}")
        if codigo_cita:
            try:
                cita = Cita.objects.get(id_cita=codigo_cita)
                data = {
                    "paciente": f"{cita.id_paciente.user.first_name} {cita.id_paciente.user.last_name}",
                    "medico": f"{cita.id_medico.user.first_name} {cita.id_medico.user.last_name}"
                }
                return JsonResponse({"success": True, "data": data})
            except Cita.DoesNotExist:
                return JsonResponse({"success": False, "error": "Cita no encontrada"})
        
        return JsonResponse({"success": False, "error": "Código no proporcionado"})
    
    return JsonResponse({"success": False, "error": "Metodo no permitido"})

@login_required
def crear_registro(request):
    if request.method == "POST":
        # Obtener los datos del formulario
        motivo = request.POST.get('motivo')
        observaciones = request.POST.get('observaciones')
        id_cita = request.POST.get('codigo_cita')
        id_enfermedad = request.POST.get('id_enfermedad')
        tratamientos_seleccionados = request.POST.getlist('tratamientos')
        productos_recomendados = request.POST.getlist('productos')

        print(f"Motivo: {motivo}")
        print(f"Observaciones: {observaciones}")
        print(f"ID Cita: {id_cita}")
        print(f"ID Enfermedad: {id_enfermedad}")
        print(f"Tratamientos seleccionados: {tratamientos_seleccionados}")
        print(f"Productos recomendados: {productos_recomendados}")

        #return HttpResponse("Cita creada")

        # Validación de los campos
        if motivo and observaciones and id_cita and id_enfermedad:
            # Crear el registro
            try:
                registro = Registro(
                    motivo=motivo,
                    observaciones=observaciones,
                    id_cita_id=id_cita,
                    id_enfermedad_id=id_enfermedad
                )
                registro.save()

                # Asociar tratamientos al registro
                for tratamiento_id in tratamientos_seleccionados:
                    RegistroTratamiento.objects.create(
                        id_registro=registro,
                        id_tratamiento_id=tratamiento_id
                    )

                # Asociar productos al registro
                for producto_id in productos_recomendados:
                    RegistroProducto.objects.create(
                        id_registro=registro,
                        id_producto_id=producto_id
                    )

                messages.success(request, '¡Registro creado exitosamente!')
                return redirect('inicio')  # Redirigir para evitar reenvío de formulario
            except Exception as e:
                messages.error(request, f'Error al crear el registro: {e}')
        else:
            messages.error(request, 'Todos los campos son obligatorios.')


@login_required
def tratamientos_por_enfermedad(request):
    enfermedad_id = request.GET.get('id_enfermedad')
    if enfermedad_id:
        tratamientos = TratamientoEnfermedad.objects.filter(id_enfermedad=enfermedad_id).select_related('id_tratamiento')

        data = [
            {
                "id_tratamiento": t.id_tratamiento.id_tratamiento,
                "nombre": t.id_tratamiento.nombre,
                "descripcion": t.id_tratamiento.descripcion
            }
            for t in tratamientos
        ]
        
        return JsonResponse({"tratamientos": data})
    
    return JsonResponse({"error": "Enfermedad no encontrada"}, status=400)